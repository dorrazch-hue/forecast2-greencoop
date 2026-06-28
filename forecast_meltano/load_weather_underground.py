import psycopg2
import os
from openpyxl import load_workbook
from datetime import datetime

conn = psycopg2.connect(
    host="localhost",
    port=5432,
    database="forecast_db",
    user="forecast_user",
    password=os.environ.get("LOCAL_DB_PASSWORD", "forecast_pass")
)
cur = conn.cursor()

cur.execute("CREATE SCHEMA IF NOT EXISTS raw_weather_underground;")
for table in ["la_madeleine", "ichtegem"]:
    cur.execute("DROP TABLE IF EXISTS raw_weather_underground." + table + " CASCADE;")
    cur.execute(
        "CREATE TABLE raw_weather_underground." + table + " ("
        "observed_at TIMESTAMP, "
        "temperature_celsius VARCHAR, "
        "dew_point_celsius VARCHAR, "
        "humidity_pct VARCHAR, "
        "wind_direction VARCHAR, "
        "wind_speed_kmh VARCHAR, "
        "wind_gust_kmh VARCHAR, "
        "pressure_hpa VARCHAR, "
        "precip_rate_mm VARCHAR, "
        "precip_accum_mm VARCHAR, "
        "uv_index VARCHAR, "
        "solar_wm2 VARCHAR"
        ");"
    )
conn.commit()

def clean(val):
    if val is None:
        return None
    return str(val).replace("\xa0", "").strip()

def f_to_c(val):
    s = clean(val)
    if not s:
        return None
    try:
        return str(round((float(s.replace("°F", "").strip()) - 32) * 5 / 9, 2))
    except Exception:
        return None

def mph_to_kmh(val):
    s = clean(val)
    if not s:
        return None
    try:
        return str(round(float(s.replace("mph", "").strip()) * 1.60934, 2))
    except Exception:
        return None

def inhg_to_hpa(val):
    s = clean(val)
    if not s:
        return None
    try:
        return str(round(float(s.replace("in", "").strip()) * 33.8639, 2))
    except Exception:
        return None

def in_to_mm(val):
    s = clean(val)
    if not s:
        return None
    try:
        return str(round(float(s.replace("in", "").strip()) * 25.4, 2))
    except Exception:
        return None

def make_ts(sheet_name, time_val):
    if time_val is None:
        return None
    try:
        month = int(sheet_name[:2])
        year = int("20" + sheet_name[4:6])
        if hasattr(time_val, "hour"):
            return datetime(year, month, 1, time_val.hour, time_val.minute)
        fval = float(time_val)
        total_min = round(fval * 24 * 60)
        return datetime(year, month, 1, (total_min // 60) % 24, total_min % 60)
    except Exception:
        return None

DATA_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data")
files = {
    "la_madeleine": os.path.join(DATA_DIR, "Weather_Underground_La_Madeleine_FR.xlsx"),
    "ichtegem": os.path.join(DATA_DIR, "Weather_Underground_Ichtegem_BE.xlsx"),
}

total = 0
for table, filepath in files.items():
    if not os.path.exists(filepath):
        print(table + ": fichier non trouve (" + filepath + ")")
        print("  -> Placez le fichier Excel dans forecast_meltano/data/")
        continue
    wb = load_workbook(filepath, read_only=True)
    count = 0
    skipped = 0
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        for row in rows[1:]:
            if not row or row[0] is None:
                skipped += 1
                continue
            ts = make_ts(sheet_name, row[0])
            if ts is None:
                skipped += 1
                continue
            humidity = clean(row[3])
            if humidity:
                humidity = humidity.replace("%", "").strip()
            solar = clean(row[11])
            if solar:
                solar = solar.replace("w/m2", "").strip()
            cur.execute(
                "INSERT INTO raw_weather_underground." + table + " "
                "(observed_at, temperature_celsius, dew_point_celsius, humidity_pct, "
                "wind_direction, wind_speed_kmh, wind_gust_kmh, pressure_hpa, "
                "precip_rate_mm, precip_accum_mm, uv_index, solar_wm2) "
                "VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)",
                (
                    ts,
                    f_to_c(row[1]),
                    f_to_c(row[2]),
                    humidity,
                    clean(row[4]),
                    mph_to_kmh(row[5]),
                    mph_to_kmh(row[6]),
                    inhg_to_hpa(row[7]),
                    in_to_mm(row[8]),
                    in_to_mm(row[9]),
                    clean(row[10]),
                    solar,
                )
            )
            count += 1
    conn.commit()
    print(table + ": " + str(count) + " releves charges (" + str(skipped) + " ignores)")
    total += count

print("Total: " + str(total) + " releves charges!")
cur.close()
conn.close()